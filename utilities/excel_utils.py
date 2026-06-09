import openpyxl

@staticmethod
def get_row_count_xl(xl_file_path,sheet):
    workbook= openpyxl.load_workbook(xl_file_path)
    sheet=workbook[sheet]
    return  sheet.max_row
    
@staticmethod
def read_data(xl_file_path,sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(xl_file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num, column_num).value